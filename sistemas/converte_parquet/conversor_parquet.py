from tkinter import ttk, filedialog, messagebox
import os
import pandas as pd
import xmltodict
import style
import utils as u
import threading
import time

interromper = False  # variável global de controle


def criar_frame_conversor_parquet(parent, btn_voltar=None):
    frame = ttk.Frame(parent, padding=10)
    conteudo_frame = ttk.Frame(frame)
    conteudo_frame.pack(fill="both", expand=True)

    arquivos_selecionados = []

    # ---------------- LOGS ----------------
    def log(msg):
        logs_widget.after(0, lambda: u.print_log(logs_widget, msg))

    # ---------------- SELECIONAR ARQUIVO ----------------
    def selecionar_arquivo():
        file = filedialog.askopenfilename(
            title="Selecione o arquivo",
            filetypes=[("Arquivos suportados", "*.xml *.csv *.json *.xlsx")]
        )
        if file:
            arquivos_selecionados.clear()
            arquivos_selecionados.append(file)
            log(f"Arquivo selecionado: {file}")
            label_progresso.config(text="Arquivo pronto para conversão.")

    ttk.Button(conteudo_frame, text="Selecionar arquivo", command=selecionar_arquivo).pack(pady=5)

    # ---------------- ELEMENTO XML ----------------
    ttk.Label(conteudo_frame, text="Elemento raiz do XML (ex: livraria/livro):").pack(pady=5)
    entrada_elemento_xml = ttk.Entry(conteudo_frame, width=50)
    entrada_elemento_xml.pack(pady=5)

    # ---------------- LABEL DE PROGRESSO ----------------
    label_progresso = ttk.Label(conteudo_frame, text="", font=("Segoe UI", 10, "italic"))
    label_progresso.pack(pady=5)

    # ---------------- LOGS ----------------
    logs_frame = ttk.Frame(conteudo_frame)
    logs_frame.pack(fill="both", expand=True, pady=5)
    logs_widget = style.criar_logs_widget(logs_frame, width=90, height=10)
    logs_widget.pack(fill="both", expand=True)

    # ---------------- CONVERSÃO ----------------
    def converter_para_parquet():
        global interromper
        interromper = False

        if not arquivos_selecionados:
            messagebox.showwarning("Aviso", "Nenhum arquivo selecionado.")
            return

        caminho_saida = filedialog.asksaveasfilename(
            defaultextension=".parquet",
            filetypes=[("Parquet", "*.parquet")]
        )
        if not caminho_saida:
            return

        elemento_xml = entrada_elemento_xml.get().strip()
        arquivo = arquivos_selecionados[0]
        ext = os.path.splitext(arquivo)[1].lower()

        log(f"Iniciando conversão: {arquivo}")
        logs_widget.after(0, lambda: label_progresso.config(text="Carregando dados..."))

        try:
            t0 = time.time()

            # função para atualizar o label de progresso
            def atualizar_progresso(linhas_lidas, total_estimado=None):
                if total_estimado:
                    perc = (linhas_lidas / total_estimado) * 100
                    label_text = f"Lidas {linhas_lidas:,}/{total_estimado:,} linhas ({perc:.1f}%)"
                else:
                    label_text = f"Lidas {linhas_lidas:,} linhas..."
                label_progresso.after(0, lambda: label_progresso.config(text=label_text))

            # ------------------- LEITURA DO ARQUIVO -------------------
            df_total = None
            linhas_lidas = 0

            if ext == ".csv":
                # leitura em chunks (não trava)
                for chunk in pd.read_csv(arquivo, chunksize=5000):
                    if interromper:
                        log("Conversão interrompida pelo usuário.")
                        label_progresso.after(0, lambda: label_progresso.config(text="Conversão interrompida."))
                        return
                    linhas_lidas += len(chunk)
                    atualizar_progresso(linhas_lidas)
                    if df_total is None:
                        df_total = chunk
                    else:
                        df_total = pd.concat([df_total, chunk], ignore_index=True)

            elif ext == ".json":
                df_total = pd.read_json(arquivo)
                atualizar_progresso(len(df_total), len(df_total))

            elif ext in [".xls", ".xlsx"]:
                import openpyxl
                import re

                log("Lendo planilhas via openpyxl...")

                wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
                sheetnames = wb.sheetnames

                # calculo do total de linhas (aproximação: considera max_row de cada folha)
                total_linhas = sum(ws.max_row - 1 for ws in wb.worksheets if ws.max_row and ws.max_row > 1)
                if total_linhas <= 0:
                    total_linhas = 1  # evita divisão por zero
                linhas_lidas = 0
                dfs = []
                t0 = time.time()

                for nome in sheetnames:
                    if interromper:
                        log("Conversão interrompida pelo usuário.")
                        label_progresso.after(0, lambda: label_progresso.config(text="Conversão interrompida."))
                        wb.close()
                        return

                    ws = wb[nome]
                    rows_iter = ws.iter_rows(values_only=True)

                    # tenta obter o cabeçalho; se for vazio, cria nomes genéricos
                    try:
                        header = list(next(rows_iter))
                    except StopIteration:
                        # planilha vazia
                        continue

                    # normaliza cabeçalho (substitui None por col_#)
                    header = [f"col_{i}" if (h is None or (isinstance(h, str) and h.strip() == "")) else str(h) for i, h in enumerate(header)]

                    batch = []
                    batch_size_for_df = 50000  # acumula este tanto antes de transformar em DataFrame (ajustável)
                    contador_no_sheet = 0

                    for row in rows_iter:
                        if interromper:
                            log("Conversão interrompida pelo usuário.")
                            label_progresso.after(0, lambda: label_progresso.config(text="Conversão interrompida."))
                            wb.close()
                            return

                        batch.append(row)
                        linhas_lidas += 1
                        contador_no_sheet += 1

                        # Atualiza a cada 1000 linhas lidas ou quando o batch atingir tamanho
                        if contador_no_sheet % 1000 == 0 or len(batch) >= batch_size_for_df or linhas_lidas == total_linhas:
                            # converte batch em DataFrame parcial
                            df_parcial = pd.DataFrame(batch, columns=header)
                            # força colunas object para str (prevenção precoce)
                            for col in df_parcial.select_dtypes(include=["object"]).columns:
                                df_parcial[col] = df_parcial[col].astype(str)

                            dfs.append(df_parcial)
                            batch = []

                            perc = (linhas_lidas / total_linhas) * 100
                            tempo_passado = time.time() - t0
                            vel = linhas_lidas / max(tempo_passado, 0.001)
                            restante = (total_linhas - linhas_lidas) / vel if vel > 0 else 0
                            label_text = f"Lidas {linhas_lidas:,}/{total_linhas:,} ({perc:.1f}%) — ETA {restante/60:.1f} min"
                            label_progresso.after(0, lambda txt=label_text: label_progresso.config(text=txt))

                    # se sobrou algo no batch após fim da folha
                    if batch:
                        df_parcial = pd.DataFrame(batch, columns=header)
                        for col in df_parcial.select_dtypes(include=["object"]).columns:
                            df_parcial[col] = df_parcial[col].astype(str)
                        dfs.append(df_parcial)
                        batch = []

                    log(f"Aba '{nome}' processada; linhas lidas até agora: {linhas_lidas:,}.")

                wb.close()

                # concatena tudo sem tentar manter tudo em memória por períodos longos
                if not dfs:
                    df_total = pd.DataFrame()  # evita erro caso não haja dados
                else:
                    df_total = pd.concat(dfs, ignore_index=True)

                # correções gerais: transforma colunas object em str (para evitar erros de inferência)
                for col in df_total.select_dtypes(include=["object"]).columns:
                    df_total[col] = df_total[col].astype(str)

                # tenta salvar em loop e, se houver erro de conversão de coluna específica, força ela para str e tenta de novo
                while True:
                    try:
                        df_total.to_parquet(caminho_saida, engine="pyarrow", index=False, compression="snappy")
                        break
                    except Exception as e:
                        msg = str(e)
                        # tenta extrair nome da coluna do erro (ex.: "Conversion failed for column 'num_medidor' with type object")
                        col_match = re.search(r"column '([^']+)'", msg)
                        if col_match:
                            col_name = col_match.group(1)
                            log(f"Forçando coluna '{col_name}' como texto (str) devido a erro de conversão...")
                            # forçar a coluna para string de forma robusta
                            if col_name in df_total.columns:
                                df_total[col_name] = df_total[col_name].apply(lambda x: "" if x is None else str(x))
                            else:
                                # se não existe exatamente assim (às vezes index/escape), tenta encontrar coluna parecida
                                possibles = [c for c in df_total.columns if col_name.lower() in c.lower()]
                                if possibles:
                                    c0 = possibles[0]
                                    log(f"Coluna exata não encontrada, aplicando correção em '{c0}'")
                                    df_total[c0] = df_total[c0].apply(lambda x: "" if x is None else str(x))
                                else:
                                    # se não houver correspondência, re-lança o erro
                                    raise e
                        else:
                            # se não foi possível identificar a coluna, relança o erro
                            raise e

            elif ext == ".xml":
                if not elemento_xml:
                    messagebox.showerror("Erro", "Para XML é necessário informar o elemento raiz.")
                    return
                with open(arquivo, 'r', encoding='utf-8') as f:
                    xml_dict = xmltodict.parse(f.read())
                data_list = xml_dict
                for tag in elemento_xml.split('/'):
                    if isinstance(data_list, dict):
                        data_list = data_list.get(tag, [])
                    else:
                        data_list = []
                if not isinstance(data_list, list):
                    data_list = [data_list]
                df_total = pd.DataFrame(data_list)
                atualizar_progresso(len(df_total), len(df_total))

            else:
                messagebox.showerror("Erro", f"Tipo de arquivo não suportado: {ext}")
                return

            # ------------------- SALVAMENTO -------------------
            log("Salvando arquivo Parquet...")
            label_progresso.after(0, lambda: label_progresso.config(text="Salvando Parquet..."))
            df_total.to_parquet(caminho_saida, engine="pyarrow", index=False, compression="snappy")

            tempo_total = time.time() - t0
            log(f"Conversão concluída em {tempo_total:.1f}s: {caminho_saida}")
            label_progresso.after(0, lambda: label_progresso.config(
                text=f"Conversão concluída em {tempo_total:.1f}s — {len(df_total):,} linhas.")
            )
            messagebox.showinfo("Sucesso", f"Parquet gerado com sucesso:\n{caminho_saida}")

        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro:\n{str(e)}")
            log(f"Erro: {str(e)}")

    # ---------------- THREADING E BOTÕES ----------------
    botoes_frame = ttk.Frame(conteudo_frame)
    botoes_frame.pack(pady=10)

    btn_converter = ttk.Button(botoes_frame, text="Converter para Parquet")
    btn_converter.pack(side="left", padx=5)

    btn_interromper = ttk.Button(botoes_frame, text="Interromper conversão", state='disabled')
    btn_interromper.pack(side="left", padx=5)

    def thread_converter():
        global interromper
        interromper = False
        btn_converter.config(state="disabled")
        btn_interromper.config(state="normal")

        def run():
            try:
                converter_para_parquet()
            finally:
                btn_converter.config(state="normal")
                btn_interromper.config(state="disabled")

        threading.Thread(target=run).start()

    def interromper_conversao():
        global interromper
        interromper = True
        label_progresso.config(text="Solicitando interrupção...")
        log("Solicitando interrupção...")

    btn_converter.config(command=thread_converter)
    btn_interromper.config(command=interromper_conversao)

    style.aplicar_estilo(frame)
    return frame, logs_widget, None
