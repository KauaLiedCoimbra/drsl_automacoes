import win32com.client
import pandas as pd
import utils as u
import re
import os
import json
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule

interrompido = False

# ------------------------------------------
# Importação JSON com relação RExNOMExE-MAIL
# ------------------------------------------
def caminho_recurso(relativo):
    # Caminho dentro do exe ou em desenvolvimento
    if getattr(sys, 'frozen', False):
        return os.path.join(sys._MEIPASS, relativo)
    return os.path.join(os.path.abspath("."), relativo)

# Carrega o JSON embutido
with open(caminho_recurso("re_nome_email.json"), "r", encoding="utf-8") as f:
    re_nomes = json.load(f)  # Será um dicionário {RE: {"Nome": ..., "E-mail": ...}}
# ------------------------------------------
# Extração dados da planilha origina
# ------------------------------------------
def extrair_dados_planilha(caminho_planilha, print_log, caminho_saida="dados_coletados.xlsx"):
    try:
        if isinstance(caminho_planilha, tuple):
            caminho, aba = caminho_planilha
            df = pd.read_excel(caminho, sheet_name=aba)
        else:
            df = pd.read_excel(caminho_planilha)
    except Exception as e:
        print_log(f"❌ Erro ao abrir a planilha {caminho_planilha}: {e}")
        return None

    # Aplicar filtros
    df_filtrado = df[
    df['EMPRESA'].str.strip().isin(['D008', 'D009']) &
    df['CÓD DO ERRO'].fillna(0).astype(int).isin([63]) &
    df['DESCRIÇÃO TIPO'].str.strip().isin([
        'Alta/Media Tensão: Optante',
        'Monômia',
        'Monômia Tarifa Branca'
    ])
]

    if df_filtrado.empty:
        print_log("⚠ Nenhum registro após aplicar os filtros.")
        return None

    # Listas para armazenar os dados extraídos
    instalacoes = []
    contratos = []
    motivos = []

    # Expressões regulares para extrair contrato e motivo
    contrato_regex = r'Contrato (\d+) bloqueado'
    motivo_regex = r'motivo bloqueio cálculo (\d+)'

    for _, row in df_filtrado.iterrows():
        instalacao = row['INSTALAÇÃO']
        descricao = str(row['DESCRIÇÃO DO ERRO'])

        # Extrair contrato e motivo
        contrato_match = re.search(contrato_regex, descricao)
        motivo_match = re.search(motivo_regex, descricao)

        contrato = contrato_match.group(1) if contrato_match else ''
        motivo = motivo_match.group(1) if motivo_match else ''

        # Armazenar
        instalacoes.append(instalacao)
        contratos.append(contrato)
        motivos.append(motivo)

    # Criar novo DataFrame
    df_saida = pd.DataFrame({
        'Instalação': instalacoes,
        'Contrato': contratos,
        'Motivo': motivos,
    })

    # Salvar planilha de saída
    df_saida.to_excel(caminho_saida, sheet_name="Dados", index=False)
    print_log(f"✅ Dados extraídos e salvos em '{caminho_saida}'")

    return caminho_saida, len(df_saida)
# ------------------------------------------
# Tratamento planilha
# ------------------------------------------
def tratar_planilha(caminho_planilha, print_log=print):
    """
    Trata a planilha adicionando colunas, formatações condicionais e estilos.
    """
    wb = load_workbook(caminho_planilha)
    
    # Aba Coleta
    if "Coleta" not in wb.sheetnames:
        print_log("❌ Aba 'Coleta' não encontrada.")
        return
    ws = wb["Coleta"]

    # Aba Encontrados
    if "Encontrados" not in wb.sheetnames:
        ws_encontrados = wb.create_sheet("Encontrados")
        ws_encontrados.cell(row=1, column=1, value="Instalação")
        ws_encontrados["A1"].font = Font(name="Calibri", size=14, bold=True)
        ws_encontrados["A1"].border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws_encontrados["A1"].alignment = Alignment(horizontal="center", vertical="center")
        print_log("✅ Aba 'Encontrados' criada vazia.")

    # -----------------------------
    # Bordas laterais e alinhamento
    # -----------------------------
    borda_lateral = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000")
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            # borda esquerda/direita
            if cell.border is None:
                cell.border = borda_lateral
            else:
                cell.border = Border(
                    left=Side(style='thin', color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
            # centraliza
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # -----------------------------
    # Formata cabeçalhos e dados existentes
    # -----------------------------
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = Font(name="Calibri", size=14, bold=True)
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).font = Font(name="Calibri", size=11)

    ultima_coluna = ws.max_column

    # -----------------------------
    # Coluna 'Encontrado?'
    # -----------------------------
    col_encontrado = ultima_coluna + 1
    ws.cell(row=1, column=col_encontrado, value="Encontrado?")
    ws.cell(row=1, column=col_encontrado).font = Font(name="Calibri", size=14, bold=True)
    ws.cell(row=1, column=col_encontrado).border = borda_lateral
    ws.cell(row=1, column=col_encontrado).alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, ws.max_row + 1):
        ws.cell(
            row=row,
            column=col_encontrado,
            value=f'=IF(ISERROR(VLOOKUP(A{row},Encontrados!A:A,1,FALSE)),"Não","Sim")'
        )
        ws.cell(row=row, column=col_encontrado).font = Font(name="Calibri", size=11)
        ws.cell(row=row, column=col_encontrado).alignment = Alignment(horizontal="center", vertical="center")

    # -----------------------------
    # Coluna 'Bloco'
    # -----------------------------
    col_bloco = col_encontrado + 1
    ws.cell(row=1, column=col_bloco, value="Bloco")
    ws.cell(row=1, column=col_bloco).font = Font(name="Calibri", size=14, bold=True)
    ws.cell(row=1, column=col_bloco).border = borda_lateral
    ws.cell(row=1, column=col_bloco).alignment = Alignment(horizontal="center", vertical="center")
    letra_bloco = ws.cell(row=1, column=col_bloco).column_letter

    ws.cell(row=2, column=col_bloco, value=1).font = Font(name="Calibri", size=11)
    ws.cell(row=2, column=col_bloco).alignment = Alignment(horizontal="center", vertical="center")

    for row in range(3, ws.max_row + 1):
        ws.cell(
            row=row,
            column=col_bloco,
            value=f'=IF(A{row}<>A{row-1},{letra_bloco}{row-1}+1,{letra_bloco}{row-1})'
        )
        ws.cell(row=row, column=col_bloco).font = Font(name="Calibri", size=11)
        ws.cell(row=row, column=col_bloco).alignment = Alignment(horizontal="center", vertical="center")

    print_log("✅ Colunas 'Encontrado?' e 'Bloco' adicionadas e centralizadas.")

    # -----------------------------
    # Borda inferior condicional quando instalação muda
    # -----------------------------
    borda_condicional = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    ws.conditional_formatting.add(
        "$A$2:$K$1048576",
        FormulaRule(formula=["=$A3<>$A2"], border=borda_condicional)
    )

    # -----------------------------
    # Formatações condicionais adicionais para 'Bloco'
    # -----------------------------
    fill1 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill2 = PatternFill(start_color="B0E5FA", end_color="B0E5FA", fill_type="solid")
    fill3 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill4 = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    ws.conditional_formatting.add(
        "$A$2:$K$1048576",
        FormulaRule(formula=["AND(ISODD($M2),ISODD(ROW()-ROW($A$2)+1))"], fill=fill1)
    )
    ws.conditional_formatting.add(
        "$A$2:$K$1048576",
        FormulaRule(formula=["AND(ISEVEN($M2),ISODD(ROW()-ROW($A$2)+1))"], fill=fill2)
    )
    ws.conditional_formatting.add(
        "$A$2:$K$7200",
        FormulaRule(formula=["ISODD($M2)"], fill=fill3)
    )
    ws.conditional_formatting.add(
        "$A$2:$K$7200",
        FormulaRule(formula=["ISEVEN($M2)"], fill=fill4)
    )

    # -----------------------------
    # Ajusta largura das colunas automaticamente
    # -----------------------------
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    # -----------------------------
    # Ajusta cor de fundo e fonte das colunas de dados adicionais
    # -----------------------------
    fill_preto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    font_branco = Font(color="FFFFFF", name="Calibri", size=11)

    # Coluna L = 12
    col_inicio = 12  # L
    linha_final = ws.max_row  # até a última linha existente (ou defina outro valor se quiser preencher mais)

    for col in range(col_inicio, ws.max_column + 1):
        for row in range(1, linha_final + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill_preto
            cell.font = font_branco
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # -----------------------------
    # Ajusta ordem das planilhas
    # -----------------------------
    if "Dados" in wb.sheetnames:
        ws_dados = wb["Dados"]
        wb._sheets.remove(ws_dados)  # Remove da posição atual
        wb._sheets.append(ws_dados)  # Adiciona no final
    # -----------------------------
    # Salva alterações
    # -----------------------------
    wb.save(caminho_planilha)
    print_log(f"🏁 Planilha tratada e salva: {caminho_planilha}")
# ------------------------------------------
# Processamento SAP
# ------------------------------------------
def executar_logs_bloqueio(caminho_filtrado=None, print_log=print, atualizar_progresso=None):
    """
    Atualiza a planilha existente criando uma aba 'Coleta' com todas as informações.
    Recebe a planilha com Instalação, Contrato e Motivo preenchidos.
    """
    global interrompido
    todos_registros = []

    print_log("🔄 Iniciando processamento dos logs de bloqueio...")

    if caminho_filtrado is None or not os.path.exists(caminho_filtrado):
        print_log("❌ Nenhum dado para processar após extração.")
        return None

    # --- Preparar DataFrame ---
    df = pd.read_excel(caminho_filtrado)
    for col in ['Instalação', 'Contrato', 'Motivo']:
        df[col] = df[col].apply(lambda x: str(int(x)) if pd.notna(x) and isinstance(x, float) else str(x).strip())

    # Conexão SAP
    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        session = SapGuiAuto.GetScriptingEngine.Children(0).Children(0)
    except Exception as e:
        print_log(f"❌ Erro ao conectar ao SAP: {e}")
        return

    # Maximiza e acessa ES21
    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/tbar[0]/okcd").text = "es21"
    session.findById("wnd[0]").sendVKey(0)

    total_contratos = len(df)
    scroll = session.findById("wnd[0]/usr").verticalScrollbar

    try:
        for index, row in df.iterrows():
            instalacao = row['Instalação']
            contrato = row['Contrato']
            motivo = row['Motivo'].zfill(2)

            print_log(f'🔍 Processando contrato {contrato}... Motivo: {motivo}')

            if interrompido:
                print_log(f"⚠ Execução interrompida pelo usuário.")
                break

            # Pesquisa contrato
            session.findById("wnd[0]/usr/ctxtEVERD-VERTRAG").text = contrato
            session.findById("wnd[0]/usr/ctxtEVERD-VERTRAG").caretPosition = len(contrato)
            session.findById("wnd[0]").sendVKey(0)
            session.findById("wnd[0]").sendVKey(19)
            session.findById("wnd[0]").sendVKey(47)

            motivo_encontrado = False
            data_atual = None
            re_atual = None
            val_antigo = ""
            val_novo = ""
            registros = []

            # Loop para ler elementos
            while not motivo_encontrado:
                if interrompido:
                    break

                # Mapeia elementos
                usr = session.findById("wnd[0]/usr")
                todos = []
                for i in range(usr.Children.Count):
                    child = usr.Children.Item(i)
                    try:
                        texto = child.Text if hasattr(child, "Text") else ""
                        if texto:
                            todos.append({"texto": texto, "top": child.Top, "left": child.Left})
                    except Exception:
                        continue

                for i, elem in enumerate(todos):
                    if interrompido:
                        break
                    texto = elem["texto"]

                    # Quando encontrar uma nova data, atualiza o RE e a data em uso
                    if u.is_data(texto) and not (i > 0 and todos[i - 1]["texto"] in ["Val.antigo:", "Val.novo:"]):
                        if texto != data_atual:
                            data_atual = texto
                            re_atual = str(todos[i - 1]["texto"] if i > 0 else "").strip().lstrip("0")
                            re_nome = re_nomes.get(re_atual, {}).get("nome", "Não encontrado")
                            re_email = re_nomes.get(re_atual, {}).get("email", "Não encontrado")
                        
                    # captura pares de valores dentro da mesma data
                    if texto == "Val.antigo:" and i + 1 < len(todos):
                        val_antigo = todos[i + 1]["texto"]
                        if val_antigo == "Val.novo:":
                            val_antigo = ""

                        # busca o próximo "Val.novo:" dentro da mesma data
                        val_novo = ""
                        if val_antigo == "":
                            j = i + 1
                        else:
                            j = i + 2
                        while j < len(todos):
                            if interrompido:
                                break
                            prox_texto = todos[j]["texto"]
                            # se achou nova data, interrompe — o próximo par pertence à próxima data
                            if u.is_data(prox_texto):
                                break
                            if prox_texto == "Val.novo:" and j + 1 < len(todos):
                                if todos[j + 1]["texto"] != "5":
                                    val_novo = todos[j + 1]["texto"]
                                    if val_novo == motivo:
                                        motivo_encontrado = True
                                break
                            j += 1

                        linha_nova = {
                            "Instalação": instalacao,
                            "Contrato": contrato,
                            "Motivo": motivo,
                            "RE": re_atual,
                            "Data": data_atual,
                            "Ano": u.extrair_ano(data_atual),
                            "Mês": u.extrair_mes(data_atual),
                            "Val.antigo": val_antigo or "",
                            "Val.novo": val_novo or "",
                            "Nome": re_nome,
                            "E-mail": re_email,
                        }
                        registros.append(linha_nova)

                        if motivo_encontrado:
                            break

                # Se não encontrou o motivo, faz o scroll e continua lendo
                if not motivo_encontrado:
                    if scroll.position >= scroll.maximum:
                        print_log(f"⚠️ Contrato {contrato}: final da tela atingido. Motivo não encontrado.")
                        registros = registros[-10:]
                        break
                    session.findById("wnd[0]").sendVKey(82)

            # Adiciona linha no DataFrame de coleta
            if registros:
                todos_registros.extend(registros)  # mantém a lógica antiga
                print_log(f"✅ Contrato {contrato} processado. Restam {total_contratos - (index + 1)} contratos.")

                if atualizar_progresso:
                    atualizar_progresso(1)

            try:
                session.StartTransaction("ES21")
            except Exception:
                pass

    except Exception as e:
        print_log(f"❌ Ocorreu um erro: {e}")

    # Cria DataFrame a partir de todos_registros
    if todos_registros:
        df_coleta = pd.DataFrame(todos_registros)
        df_encontrados = pd.DataFrame(columns=['Instalação'])

        # Salva em uma aba nova 'Coleta'
        with pd.ExcelWriter(caminho_filtrado, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_coleta.to_excel(writer, sheet_name="Coleta", index=False)
            df_encontrados.to_excel(writer, sheet_name="Encontrados", index=False)

        tratar_planilha(caminho_filtrado, print_log)

        print_log(f'🏁 Processamento finalizado. Aba "Coleta" atualizada em: {caminho_filtrado}')